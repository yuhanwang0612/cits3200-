"""
ABDC journal rating join — CITS3200 Group 20.

Adds the ABDC rating for each journal in a publications CSV. Written against
the UNSW data but deliberately not tied to it: the journal column is detected
rather than hard-coded, so it runs on the ANU, UQ and Monash exports too.

    python abdc.py --publications output/unsw_publications.csv \
                   --abdc ABDC-JQL-2025-v2-270526.xlsx

Writes <name>_with_abdc.csv (every original column, plus the ABDC ones) and
<name>_abdc_unmatched.csv (the journals it could not rate, most frequent first
— that file is the to-do list, not an error log).

THE LIST FILE IS NOT IN THIS REPO
---------------------------------
Download it yourself from https://abdc.edu.au/abdc-journal-quality-list/. It is
free and needs no login, but the ABDC does not grant redistribution, so it is
gitignored rather than committed. One workbook holds several years as separate
sheets; --year picks one, and the default is the most recent.

WHY THERE IS NO FUZZY MATCHING BY DEFAULT
-----------------------------------------
A wrong A* is far worse than a blank. Fuzzy matching on journal titles produces
exactly that: "Journal of Accounting Research" (A*) and "Journal of Accounting
Education" (B) are three characters apart. So matching is exact, on ISSN first
and then on a normalised title. Anything that does not match is reported rather
than guessed at, and every matched row records HOW it matched in
abdc_match_type so a reviewer can audit the joins.

--fuzzy is available for working through the unmatched list by hand. It is
deliberately conservative, never overrides a real match, and tags its results
as "fuzzy" so they can be filtered out or checked.
"""

import argparse
import csv
import os
import re
import sys
from collections import Counter

import journal_match as jm

try:
    from openpyxl import load_workbook
except ImportError:                                   # pragma: no cover
    sys.exit("openpyxl is required:  python -m pip install openpyxl")


ADDED_COLUMNS = ["abdc_rating", "abdc_for_code", "abdc_matched_title",
                 "abdc_match_type", "abdc_list_year"]

# Column names to look for in the ABDC workbook and in publications CSVs.
# Both are matched case-insensitively on a normalised header, because the
# workbook's headers have changed between editions and our four scrapers do
# not all use the same field name.
ABDC_TITLE_HEADERS = ("journal title", "title", "journal name", "journal")
ABDC_RATING_HEADERS = ("rating", "abdc rating", "current rating")
ABDC_ISSN_HEADERS = ("issn", "issn print", "print issn")
# The 2025 sheet writes this as "ISSNOnline" with no space; 2022 uses
# "ISSN Online". Normalising punctuation does not join or split words, so both
# spellings have to be listed.
ABDC_ISSN_ONLINE_HEADERS = ("issn online", "issnonline", "online issn",
                            "eissn", "e-issn")
ABDC_FOR_HEADERS = ("for", "for code", "field of research", "for")


VALID_RATINGS = {"A*", "A", "B", "C"}

# Written into abdc_rating when a journal is named but is not on the ABDC list.
# Requested by the client, 12 August 2026.
UNRATED = "none"


# ---------------------------------------------------------------------------
# Normalisation and matching live in journal_match.py, shared with scimago.py.
# Re-exported here so callers and tests can keep using abdc.normalise(...) —
# and, more importantly, so a journal that ABDC matches is matched the same way
# by every other ranking source we join.
# ---------------------------------------------------------------------------
normalise = jm.normalise
normalise_issn = jm.normalise_issn
build_aliases = jm.build_aliases
fuzzy_match = jm.fuzzy_match
ISSN_RE = jm.ISSN_RE
SUBTITLE = jm.SUBTITLE


def match_journal(journal_name, issn, index, aliases=None):
    """Thin wrapper: ABDC calls the shared prefix match 'abdc-prefix'."""
    record, how = jm.match_journal(journal_name, issn, index, aliases)
    return record, ("abdc-prefix" if how == "prefix" else how)


def clean_rating(value):
    """ABDC ratings are A*, A, B, C. Anything else is not a rating.

    The workbook contains trailing spaces on almost every value ("B "), and a
    Greek capital alpha in at least one row, which looks identical to "A".
    """
    if not value:
        return None
    text = str(value).strip().upper().replace("Α", "A")
    return text if text in VALID_RATINGS else None


# ---------------------------------------------------------------------------
# Reading the ABDC workbook
# ---------------------------------------------------------------------------
def _find_column(headers, candidates):
    """Return the index of the first header matching one of `candidates`."""
    normalised = [normalise(h) for h in headers]
    for candidate in candidates:
        key = normalise(candidate)
        if key in normalised:
            return normalised.index(key)
    # Fall back to a header that starts or ends with the candidate. The rating
    # column is named for its edition — "2025 Rating", "2019 Rating" — so an
    # exact match on "rating" alone never fires, and each new list would break
    # this module until someone added another literal.
    for candidate in candidates:
        key = normalise(candidate)
        for i, header in enumerate(normalised):
            if header.startswith(key) or header.endswith(key):
                return i
    return None


def _header_row(rows, candidates, limit=12):
    """The ABDC sheets carry a title banner above the real header row."""
    for index, row in enumerate(rows[:limit]):
        values = [c for c in row if c]
        if len(values) >= 3 and _find_column(row, candidates) is not None:
            return index
    return None


def load_abdc(path, year=None):
    """Read the ABDC workbook into lookup tables.

    Returns (index, chosen_sheet_name). `index` maps both normalised titles and
    ISSNs to {'rating', 'for_code', 'title'}.
    """
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = workbook.sheetnames

    if year:
        matching = [s for s in sheets if str(year) in s]
        if not matching:
            raise SystemExit(
                f"No sheet for {year} in {os.path.basename(path)}. "
                f"Sheets present: {', '.join(sheets)}")
        sheet_name = matching[0]
    else:
        # Default to the most recent year named in a sheet title. One workbook
        # holds several editions (2025, 2019, 2016, …) as separate sheets.
        years = [(max((int(y) for y in re.findall(r"\b((?:19|20)\d{2})\b", s)),
                      default=0), s) for s in sheets]
        sheet_name = max(years)[1] if max(years)[0] else sheets[0]

    rows = [list(r) for r in workbook[sheet_name].iter_rows(values_only=True)]
    workbook.close()

    header_index = _header_row(rows, ABDC_TITLE_HEADERS)
    if header_index is None:
        raise SystemExit(
            f"Could not find a journal-title column in sheet '{sheet_name}'. "
            "The workbook layout may have changed — check the header row.")

    headers = rows[header_index]
    col_title = _find_column(headers, ABDC_TITLE_HEADERS)
    col_rating = _find_column(headers, ABDC_RATING_HEADERS)
    col_issn = _find_column(headers, ABDC_ISSN_HEADERS)
    col_issn_online = _find_column(headers, ABDC_ISSN_ONLINE_HEADERS)
    col_for = _find_column(headers, ABDC_FOR_HEADERS)

    if col_rating is None:
        raise SystemExit(
            f"Found titles but no rating column in sheet '{sheet_name}'. "
            f"Headers seen: {[h for h in headers if h]}")

    index, skipped = {}, 0
    for row in rows[header_index + 1:]:
        if col_title >= len(row):
            continue
        title = row[col_title]
        rating = clean_rating(row[col_rating] if col_rating < len(row) else None)
        if not title or not rating:
            skipped += 1
            continue
        record = {
            "rating": rating,
            "for_code": (row[col_for] if col_for is not None
                         and col_for < len(row) else None),
            "title": str(title).strip(),
            # Kept on the record, not just used as a lookup key: our scrapers
            # mostly cannot get an ISSN from the university websites, so the
            # reference lists are where it comes from. journals.py reads these.
            "issn": (normalise_issn(row[col_issn])
                     if col_issn is not None and col_issn < len(row) else None),
            "issn_online": (normalise_issn(row[col_issn_online])
                            if col_issn_online is not None
                            and col_issn_online < len(row) else None),
        }
        key = normalise(title)
        if key:
            index.setdefault(key, record)
        for issn in (record["issn"], record["issn_online"]):
            if issn:
                index.setdefault(issn, record)

    if not index:
        raise SystemExit(f"Sheet '{sheet_name}' produced no usable rows.")
    return index, sheet_name, skipped


# ---------------------------------------------------------------------------
def enrich(publications_path, abdc_path, year=None, use_fuzzy=False,
           journal_column=None):
    index, sheet_name, skipped = load_abdc(abdc_path, year)
    aliases = build_aliases(index)
    list_year = (re.search(r"\b((?:19|20)\d{2})\b", sheet_name) or [None])[0] \
        if re.search(r"\b((?:19|20)\d{2})\b", sheet_name) else sheet_name
    print(f"ABDC list: sheet '{sheet_name}', {len(index)} lookup keys, "
          f"{len(aliases)} unambiguous subtitle aliases "
          f"({skipped} rows without a usable title+rating skipped)")

    rows, fieldnames = jm.read_publications(publications_path)
    column, issn_column = jm.find_columns(fieldnames, journal_column)
    print(f"Publications: {len(rows)} rows, matching on '{column}'"
          + (f" and '{issn_column}'" if issn_column else ""))

    counts = Counter()
    unmatched = Counter()
    for row in rows:
        journal = row.get(column)
        record, how = match_journal(journal,
                                    row.get(issn_column) if issn_column else None,
                                    index, aliases)
        if record is None and use_fuzzy:
            record, how = fuzzy_match(journal, index)
        if record is None:
            counts["unmatched"] += 1
            row.update({c: None for c in ADDED_COLUMNS})
            if journal and journal.strip():
                unmatched[journal.strip()] += 1
                # The client asked on 12 August for publications that are not in
                # ABDC to carry a rating of "none" rather than an empty cell.
                # A journal that exists but is unranked is a finding; a blank
                # cell reads as "we didn't check". Rows with no journal at all
                # are left blank, because there was nothing to rate.
                row["abdc_rating"] = UNRATED
            else:
                counts["no journal name"] += 1
            continue
        counts[how] += 1
        row.update({
            "abdc_rating": record["rating"],
            "abdc_for_code": record["for_code"],
            "abdc_matched_title": record["title"],
            "abdc_match_type": how,
            "abdc_list_year": list_year,
        })

    out_path, unmatched_path = jm.write_enriched(
        publications_path, fieldnames, ADDED_COLUMNS, rows, unmatched, "with_abdc")
    jm.report(counts, unmatched, len(rows), out_path, unmatched_path,
              skip_keys={"unmatched", "no journal name"})
    return out_path, unmatched_path


def main():
    parser = argparse.ArgumentParser(
        description="Add ABDC journal ratings to a publications CSV.")
    parser.add_argument("--publications", required=True,
                        help="a *_publications.csv from any of our scrapers")
    parser.add_argument("--abdc", required=True,
                        help="the ABDC Journal Quality List .xlsx")
    parser.add_argument("--year", help="which list year to use (default: latest)")
    parser.add_argument("--journal-column",
                        help="override the journal-name column name")
    parser.add_argument("--fuzzy", action="store_true",
                        help="also attempt conservative fuzzy title matching "
                             "(results are tagged abdc_match_type=fuzzy)")
    args = parser.parse_args()
    enrich(args.publications, args.abdc, args.year, args.fuzzy,
           args.journal_column)


if __name__ == "__main__":
    main()
